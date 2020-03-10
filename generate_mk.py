import openpyxl

file_name = 'demo.md'

sota_data = openpyxl.load_workbook('sota_data.xlsx')
worksheet = sota_data.active
nrows = worksheet.max_row
ncols = worksheet.max_column

raw_data = []
for row in range(1, nrows + 1):
    temp = []
    for col in range(1, ncols + 1):
        temp.append(worksheet.cell(row, col).value)
    raw_data.append(temp)
# print(raw_data)

length = len(raw_data)
print(length)

with open(file_name, 'a+') as f:
    f.write('# AI')
    for i in range(length):
        main_category = raw_data[i][2]
        second_category = raw_data[i][6]
        third_category = raw_data[i][8]
        if main_category!='':
            f.write('\n## ' + main_category)
        if second_category!='':
            f.write('\n### ' + second_category)
        f.write('\n#### ' + third_category)
        # if main_category != '':
        #     f.write('\n## ' + main_category)
        #     if second_category != '':
        #         f.write('\n###' + second_category)
        #         f.write('\n####' + third_category)
        #     else:
        #
        # else:
        #     if second_category != '':
        #         f.write('\n###' + second_category)
        #     else:
        #         f.write('\n#### ' + third_category)
