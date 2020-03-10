import openpyxl

sota_data = openpyxl.load_workbook('sota_data.xlsx')
worksheet = sota_data.active
nrows = worksheet.max_row
ncols = worksheet.max_column

temp1 = ''
for cell in worksheet['A']:
    temp1 = cell.value
    print(temp1)
